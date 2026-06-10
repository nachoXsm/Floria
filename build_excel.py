import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from copy import copy

wb = openpyxl.load_workbook('flora_argentina_completa_v4.xlsx')
ws = wb.active

# Style templates from existing rows
FILL_ODD  = PatternFill("solid", fgColor="F0FDF4")   # light green
FILL_EVEN = PatternFill("solid", fgColor="FFFFFF")    # white

THIN = Side(border_style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FONT = Font(name="Arial", size=9)
ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Columns that should be left-aligned (text-heavy)
LEFT_COLS = {2, 3, 15, 16, 17}  # B=Nombre Vulgar, C=Científico, O=Uso, P=Compatibilidad, Q=Notas (1-indexed)

def apply_style(row_num):
    fill = FILL_ODD if row_num % 2 == 0 else FILL_EVEN  # row 2 is odd index but even fill
    # Match the template: row 2 (first data row) = light green, row 3 = white
    fill = FILL_ODD if (row_num % 2 == 0) else FILL_EVEN
    row = ws[row_num]
    for i, cell in enumerate(row, start=1):
        cell.fill = fill
        cell.border = BORDER
        cell.font = FONT
        cell.alignment = ALIGN_LEFT if i in LEFT_COLS else ALIGN

# Fix Lote 1 rows (282-438)
print("Applying formatting to Lote 1 rows (282-438)...")
for r in range(282, 439):
    apply_style(r)

# ─── LOTE 2: Frutales y Aromáticas ────────────────────────────────────────────
# Columns: N°, Nombre Vulgar, Nombre Científico, Tipo, Ubicación, Luz, Riego,
#          Floración, Perenne/Anual, Altura(m), Diámetro(m), Siembra, Región,
#          Nativa/Exótica, Uso Paisajismo, Compatibilidad Diseño, Notas,
#          Color Floración, Apto Maceta

lote2 = [
    # Frutales
    [439,"Manzano","Malus domestica","Frutal","Exterior","Alta","Moderado","Primavera","Caducifolio","3-8","3-6","Otoño","Centro/Sur","Exótica","Frutal / Jardín productivo","Huerto familiar, jardín rural","Necesita dos variedades para polinización cruzada","Blanco/Rosa","No"],
    [440,"Peral","Pyrus communis","Frutal","Exterior","Alta","Moderado","Primavera","Caducifolio","4-10","3-6","Otoño","Centro/Sur","Exótica","Frutal / Jardín productivo","Huerto familiar, diseño rural","Tolera suelos arcillosos mejor que el manzano","Blanco","No"],
    [441,"Duraznero","Prunus persica","Frutal","Exterior","Alta","Moderado","Primavera","Caducifolio","3-6","3-5","Otoño","Centro/Norte","Exótica","Frutal ornamental","Jardín mixto, huerto","Floración rosada muy ornamental en primavera","Rosa","No"],
    [442,"Ciruelo","Prunus domestica","Frutal","Exterior","Alta","Moderado","Primavera","Caducifolio","3-6","3-5","Otoño","Centro/Sur","Exótica","Frutal / Jardín productivo","Huerto familiar","Varias cultivares con frutos rojos, amarillos o verdes","Blanco/Rosa","No"],
    [443,"Cerezo","Prunus avium","Frutal","Exterior","Alta","Moderado","Primavera","Caducifolio","5-10","4-7","Otoño","Sur","Exótica","Frutal ornamental / Jardín","Parque, jardín rural","Requiere frío invernal; floración espectacular","Blanco","No"],
    [444,"Higuera","Ficus carica","Frutal","Exterior","Alta","Escaso","Verano-Otoño","Caducifolio","3-6","3-5","Primavera","Centro/Norte","Exótica","Frutal / Patio / Jardín","Jardín mediterráneo, patio","Muy resistente a la sequía una vez establecida","Sin flor visible","No"],
    [445,"Naranjo","Citrus sinensis","Frutal","Exterior","Alta","Moderado","Primavera","Perenne","3-6","3-5","Primavera","Norte/Centro","Exótica","Frutal / Ornamental","Jardín subtropical, maceta grande","Requiere clima sin heladas fuertes; fragancia intensa","Blanco","Sí"],
    [446,"Limonero","Citrus limon","Frutal","Exterior","Alta","Moderado","Todo el año","Perenne","2-5","2-4","Primavera","Norte/Centro","Exótica","Frutal / Ornamental / Aromática","Jardín subtropical, terraza","Florece varias veces al año; muy fragante","Blanco","Sí"],
    [447,"Mandarino","Citrus reticulata","Frutal","Exterior","Alta","Moderado","Primavera","Perenne","2-5","2-4","Primavera","Norte/Centro","Exótica","Frutal / Ornamental","Jardín subtropical, maceta","Fruto dulce; más resistente al frío que naranjo","Blanco","Sí"],
    [448,"Pomelo","Citrus paradisi","Frutal","Exterior","Alta","Moderado","Primavera","Perenne","4-8","4-6","Primavera","Norte","Exótica","Frutal / Jardín","Jardín subtropical","Árbol de gran porte; poco resistente al frío","Blanco","No"],
    [449,"Kiwi","Actinidia deliciosa","Trepadora frutal","Exterior","Alta","Moderado","Primavera","Caducifolio","6-9 (trepadora)","3-5","Otoño","Centro/Sur","Exótica","Frutal / Pérgola / Cerco","Pérgola, muro, jardín productivo","Dioica: necesita planta macho y hembra; vigorosa","Blanco/Crema","No"],
    [450,"Frambuesa","Rubus idaeus","Frutal","Exterior","Alta","Moderado","Primavera-Verano","Caducifolio","1-2","1","Otoño","Sur","Exótica","Frutal / Jardín productivo","Huerto, jardín rural","Requiere soporte; producción el segundo año","Blanco","Sí"],
    [451,"Arándano","Vaccinium corymbosum","Frutal","Exterior","Alta","Moderado","Primavera","Caducifolio","1-2","1","Otoño","Sur/Centro","Exótica","Frutal / Ornamental","Huerto, jardín moderno","Necesita suelo ácido pH 4.5-5.5; bello follaje otoñal","Blanco/Rosa","Sí"],
    [452,"Frutilla","Fragaria × ananassa","Frutal","Exterior","Alta","Frecuente","Primavera","Perenne baja","0.2-0.3","0.3","Otoño","Todo el país","Exótica","Frutal / Borde / Maceta","Jardín comestible, maceta, borde","Produce estolones; renovar cada 2-3 años","Blanco","Sí"],
    [453,"Mora","Morus nigra","Frutal","Exterior","Alta","Moderado","Primavera","Caducifolio","5-12","5-10","Otoño","Centro/Norte","Exótica","Frutal / Sombra / Ornamental","Parque, jardín grande","Frutos manchan; hojas caducas grandes","Sin flor vistosa","No"],
    [454,"Níspero","Eriobotrya japonica","Frutal","Exterior","Alta","Moderado","Otoño-Invierno","Perenne","4-8","4-6","Otoño","Centro/Norte","Exótica","Frutal ornamental / Jardín","Jardín mediterráneo, subtropical","Fruto a fines de invierno; hojas grandes ornamentales","Blanco/Crema","No"],
    [455,"Membrillo","Cydonia oblonga","Frutal","Exterior","Alta","Moderado","Primavera","Caducifolio","3-6","3-4","Otoño","Centro/Sur","Exótica","Frutal / Jardín rural","Huerto familiar","Muy ornamental en flor; fruto solo para cocción","Rosa/Blanco","No"],
    [456,"Granado","Punica granatum","Frutal","Exterior","Alta","Escaso","Primavera-Verano","Caducifolio","2-5","2-4","Primavera","Centro/Norte","Exótica","Frutal / Ornamental / Cerco","Jardín mediterráneo, xeriscape","Muy resistente a sequía y calor; flores naranjas vistosas","Naranja/Rojo","No"],
    [457,"Olivo","Olea europaea","Frutal","Exterior","Alta","Escaso","Primavera","Perenne","4-10","4-8","Primavera","Centro/Norte","Exótica","Frutal / Ornamental / Estructura","Jardín mediterráneo, rural","Centenario; muy resistente a sequía; fruto comestible","Amarillo/Blanco","No"],
    [458,"Palto / Aguacate","Persea americana","Frutal","Exterior","Alta","Moderado","Primavera","Perenne","5-12","5-10","Primavera","Norte","Exótica","Frutal / Sombra","Jardín subtropical","No tolera heladas; requiere 2 plantas para buena cosecha","Amarillo/Verde","No"],
    [459,"Palta (criolla)","Persea drymifolia","Frutal","Exterior","Alta","Moderado","Primavera","Perenne","4-8","3-6","Primavera","Norte","Exótica","Frutal / Ornamental","Jardín subtropical","Más chica que el aguacate común; fruto pequeño","Amarillo","No"],
    [460,"Mango","Mangifera indica","Frutal","Exterior","Alta","Moderado","Invierno-Primavera","Perenne","10-20","6-10","Primavera","Norte","Exótica","Frutal / Sombra / Ornamental","Jardín tropical grande","Solo en zonas sin heladas; árbol imponente","Amarillo/Rojo","No"],
    [461,"Uva (vid)","Vitis vinifera","Frutal trepadora","Exterior","Alta","Moderado","Primavera","Caducifolio","3-8 (trepadora)","2-4","Otoño","Centro/Norte","Exótica","Frutal / Pérgola / Enrame","Pérgola, jardín rural, patio","Requiere poda anual; múltiples cultivares","Sin flor vistosa","No"],
    [462,"Maracuyá","Passiflora edulis","Frutal trepadora","Exterior","Alta","Moderado","Verano","Perenne","3-5 (trepadora)","2-3","Primavera","Norte","Exótica","Frutal / Trepadora ornamental","Pérgola, jardín tropical","Flor muy ornamental; fruto aromático","Blanco/Violeta","No"],
    [463,"Banana / Bananero","Musa × paradisiaca","Frutal","Exterior","Alta","Frecuente","Variable","Perenne herbácea","2-4","1.5-2.5","Primavera","Norte","Exótica","Frutal / Tropical / Estructura","Jardín tropical, diseño exótico","No es árbol sino hierba gigante; requiere calor y humedad","Crema/Amarillo","No"],
    [464,"Papaya","Carica papaya","Frutal","Exterior","Alta","Moderado","Todo el año","Perenne corta vida","2-5","1-2","Primavera","Norte","Exótica","Frutal / Ornamental tropical","Jardín subtropical","De rápido crecimiento; vida útil 3-5 años","Blanco/Crema","No"],
    # Aromáticas
    [465,"Lavanda","Lavandula angustifolia","Aromática / Ornamental","Exterior","Alta","Escaso","Primavera-Verano","Perenne","0.5-1","0.5-1","Primavera","Todo el país","Exótica","Borde / Aromática / Xeriscape","Jardín mediterráneo, borde, rockery","Repele mosquitos; muy fragante; atrae polinizadores","Violeta/Lila","Sí"],
    [466,"Romero","Salvia rosmarinus","Aromática","Exterior","Alta","Escaso","Primavera","Perenne","0.5-1.5","0.5-1","Primavera","Todo el país","Exótica","Aromática / Seto bajo / Xeriscape","Jardín mediterráneo, huerto","Uso culinario y medicinal; muy resistente","Azul/Violeta","Sí"],
    [467,"Tomillo","Thymus vulgaris","Aromática","Exterior","Alta","Escaso","Primavera","Perenne","0.2-0.4","0.3-0.5","Primavera","Todo el país","Exótica","Aromática / Tapizante / Borde","Rockery, borde, jardín mediterráneo","Pisarlo libera aroma; uso culinario","Rosa/Blanco","Sí"],
    [468,"Salvia","Salvia officinalis","Aromática","Exterior","Alta","Escaso","Primavera-Verano","Perenne","0.4-0.8","0.4-0.6","Primavera","Todo el país","Exótica","Aromática / Medicinal / Borde","Jardín mediterráneo, huerto, borde","Hojas plateadas muy ornamentales; uso culinario","Violeta/Azul","Sí"],
    [469,"Menta","Mentha spicata","Aromática","Exterior/Interior","Media","Frecuente","Verano","Perenne","0.3-0.6","Invasiva","Primavera","Todo el país","Exótica","Aromática / Medicinal","Maceta (para controlar invasividad)","Invasiva en jardín; mejor en maceta; refrescante","Lila/Rosa","Sí"],
    [470,"Orégano","Origanum vulgare","Aromática","Exterior","Alta","Escaso","Verano","Perenne","0.3-0.6","0.3-0.5","Primavera","Todo el país","Exótica","Aromática / Tapizante","Jardín mediterráneo, huerto","Uso culinario; atrae mariposas","Rosa/Blanco","Sí"],
    [471,"Albahaca","Ocimum basilicum","Aromática","Exterior","Alta","Moderado","Verano","Anual","0.3-0.6","0.2-0.4","Primavera","Centro/Norte","Exótica","Aromática / Huerto","Maceta, huerto, jardín","Plantar junto a tomates; sensible al frío","Blanco","Sí"],
    [472,"Perejil","Petroselinum crispum","Aromática","Exterior/Interior","Media","Moderado","Verano","Bienal","0.2-0.4","0.2-0.3","Otoño/Primavera","Todo el país","Exótica","Aromática / Huerto","Maceta, huerto","Bienal; primer año hojas, segundo año flores","Amarillo/Verde","Sí"],
    [473,"Ciboulette","Allium schoenoprasum","Aromática","Exterior","Alta","Moderado","Primavera-Verano","Perenne","0.2-0.4","0.2-0.3","Otoño/Primavera","Todo el país","Exótica","Aromática / Borde / Huerto","Borde, maceta, huerto","Flores comestibles y ornamentales; fácil cultivo","Violeta/Rosa","Sí"],
    [474,"Estragón","Artemisia dracunculus","Aromática","Exterior","Alta","Escaso","Verano","Perenne","0.6-1","0.3-0.5","Primavera","Centro/Sur","Exótica","Aromática / Medicinal","Huerto, jardín mediterráneo","Sabor anisado; prefiere clima templado-frío","Amarillo/Verde","Sí"],
    [475,"Eneldo","Anethum graveolens","Aromática","Exterior","Alta","Moderado","Verano","Anual","0.6-1","0.3-0.5","Primavera","Centro/Sur","Exótica","Aromática / Huerto","Huerto, jardín naturalista","Se reseed solo; umbelas ornamentales","Amarillo","Sí"],
    [476,"Cilantro","Coriandrum sativum","Aromática","Exterior","Alta","Moderado","Primavera","Anual","0.3-0.6","0.2-0.3","Otoño/Primavera","Todo el país","Exótica","Aromática / Huerto","Huerto, maceta","Sensible al calor; bolting rápido en verano","Blanco","Sí"],
    [477,"Hinojo","Foeniculum vulgare","Aromática","Exterior","Alta","Escaso","Verano","Perenne","1-2","0.5-1","Primavera","Todo el país","Exótica/Naturalizada","Aromática / Medicinal / Estructural","Jardín naturalista, borde alto","Sabor anisado; plumoso ornamental; semi-invasivo","Amarillo","No"],
    [478,"Hierbabuena","Mentha × piperita","Aromática","Exterior/Interior","Media","Frecuente","Verano","Perenne","0.3-0.6","Invasiva","Primavera","Todo el país","Exótica","Aromática / Medicinal","Maceta (controlar invasividad)","Cruce de menta y menta acuática; muy fragante","Lila","Sí"],
    [479,"Melisa / Toronjil","Melissa officinalis","Aromática","Exterior","Alta/Media","Moderado","Verano","Perenne","0.4-0.8","0.4-0.6","Primavera","Todo el país","Exótica","Aromática / Medicinal / Borde","Borde, jardín naturalista, maceta","Aroma cítrico suave; atrae abejas; sedante","Blanco","Sí"],
    [480,"Manzanilla","Matricaria chamomilla","Aromática / Medicinal","Exterior","Alta","Moderado","Primavera","Anual","0.2-0.5","0.2-0.3","Otoño/Primavera","Todo el país","Exótica","Medicinal / Tapizante / Borde","Jardín naturalista, borde","Flores para infusión; re-siembra sola","Blanco/Amarillo","Sí"],
    [481,"Caléndula","Calendula officinalis","Aromática / Ornamental","Exterior","Alta","Moderado","Otoño-Primavera","Anual","0.3-0.6","0.2-0.4","Otoño","Todo el país","Exótica","Borde / Medicinal / Huerto","Borde, huerto compañero, maceta","Repele plagas del huerto; flores comestibles","Naranja/Amarillo","Sí"],
    [482,"Cedrón","Aloysia citrodora","Aromática","Exterior","Alta","Moderado","Verano","Caducifolio","1-2","1-1.5","Primavera","Centro/Norte","Nativa","Aromática / Medicinal / Borde","Jardín nativo, borde, maceta","Hierba nacional argentina; aroma cítrico intenso","Blanco/Lila","Sí"],
    [483,"Burrito / Hierbabuena del campo","Lippia integrifolia","Aromática","Exterior","Alta","Escaso","Primavera-Verano","Perenne","0.5-1","0.5-0.8","Primavera","Centro/Norte","Nativa","Aromática / Medicinal","Jardín nativo, xeriscaping","Nativa argentina; aroma intenso a menta; resistente","Blanco/Amarillo","Sí"],
    [484,"Poleo","Lippia turbinata","Aromática","Exterior","Alta","Escaso","Primavera-Verano","Perenne","0.3-0.6","0.4-0.6","Primavera","Centro/Norte","Nativa","Aromática / Medicinal","Jardín nativo","Aromática nativa; uso medicinal tradicional","Blanco","Sí"],
    [485,"Peperina","Minthostachys mollis","Aromática","Exterior","Alta","Escaso","Primavera-Verano","Perenne","0.3-0.6","0.3-0.5","Primavera","Centro/Noroeste","Nativa","Aromática / Medicinal","Jardín nativo, rockery","Endémica del NOA; aroma intenso; uso en mate","Blanco/Lila","Sí"],
    [486,"Ruda","Ruta graveolens","Aromática","Exterior","Alta","Escaso","Verano","Perenne","0.5-1","0.4-0.6","Primavera","Centro/Norte","Exótica","Medicinal / Repelente / Borde","Jardín tradicional","Repele insectos; uso cultural argentino; tóxica en exceso","Amarillo","Sí"],
    [487,"Ajenjo","Artemisia absinthium","Aromática","Exterior","Alta","Escaso","Verano","Perenne","0.6-1.2","0.5-0.8","Primavera","Centro/Sur","Exótica","Medicinal / Repelente / Estructural","Jardín mediterráneo, borde plateado","Follaje plateado muy ornamental; repele plagas","Amarillo/Gris","No"],
    [488,"Santolina","Santolina chamaecyparissus","Aromática","Exterior","Alta","Escaso","Verano","Perenne","0.4-0.6","0.4-0.6","Primavera","Centro/Norte","Exótica","Aromática / Borde / Seto bajo","Jardín mediterráneo, xeriscape","Follaje plateado; repele polillas; muy ornamental","Amarillo","Sí"],
    [489,"Anís","Pimpinella anisum","Aromática","Exterior","Alta","Moderado","Verano","Anual","0.3-0.6","0.2-0.3","Primavera","Centro/Norte","Exótica","Aromática / Huerto","Huerto, jardín naturalista","Semillas para infusión y cocina; aroma intenso","Blanco","Sí"],
    [490,"Curry (planta de)","Helichrysum italicum","Aromática","Exterior","Alta","Escaso","Verano","Perenne","0.4-0.6","0.4-0.6","Primavera","Centro/Norte","Exótica","Aromática / Ornamental / Borde","Jardín mediterráneo, borde","No es el curry especiado; aroma similar; flores amarillas","Amarillo","Sí"],
    [491,"Bergamota / Monarda","Monarda didyma","Aromática / Ornamental","Exterior","Alta","Moderado","Verano","Perenne","0.6-1","0.4-0.6","Primavera","Centro/Sur","Exótica","Aromática / Borde ornamental","Jardín naturalista, borde","Flores rojas vistosas; atrae colibríes y mariposas","Rojo/Violeta","Sí"],
    [492,"Limonaria","Cymbopogon citratus","Aromática","Exterior/Interior","Alta","Moderado","Verano","Perenne","1-1.5","0.5-1","Primavera","Norte","Exótica","Aromática / Estructural / Tropical","Jardín tropical, maceta grande","Gramínea tropical; aroma cítrico intenso; no soporta heladas","Crema","Sí"],
    [493,"Valeriana","Valeriana officinalis","Aromática / Medicinal","Exterior","Alta/Media","Moderado","Primavera-Verano","Perenne","0.8-1.5","0.4-0.6","Otoño/Primavera","Centro/Sur","Exótica","Medicinal / Borde alto","Jardín naturalista, borde","Raíz sedante; flores pequeñas; atrae gatos (valerian)","Blanco/Rosa","No"],
    [494,"Ysopo","Hyssopus officinalis","Aromática","Exterior","Alta","Escaso","Verano","Perenne","0.4-0.6","0.3-0.5","Primavera","Centro/Sur","Exótica","Aromática / Borde / Medicinal","Jardín mediterráneo, borde","Atrae abejas; uso culinario y medicinal","Azul/Violeta","Sí"],
    [495,"Salvia romana","Salvia sclarea","Aromática / Ornamental","Exterior","Alta","Escaso","Primavera-Verano","Bienal","0.6-1","0.4-0.6","Otoño","Centro/Norte","Exótica","Aromática / Borde alto","Jardín mediterráneo, borde","Flores bicolores muy ornamentales; bienal","Blanco/Lila","No"],
    [496,"Lavanda dentada","Lavandula dentata","Aromática","Exterior","Alta","Escaso","Casi todo el año","Perenne","0.5-0.8","0.5-0.8","Primavera","Centro/Norte","Exótica","Aromática / Borde / Xeriscape","Jardín mediterráneo, borde","Florece más tiempo que L. angustifolia; hojas dentadas ornamentales","Violeta","Sí"],
    [497,"Canelo","Cinnamomum verum","Aromática / Árbol","Exterior","Alta","Moderado","Primavera","Perenne","5-15","3-6","Primavera","Norte","Exótica","Aromática / Sombra / Estructural","Jardín tropical, diseño exótico","Corteza y hojas aromáticas; necesita calor","Blanco/Crema","No"],
    [498,"Boldo","Peumus boldus","Aromática / Medicinal","Exterior","Alta/Media","Escaso","Invierno-Primavera","Perenne","2-6","2-4","Primavera","Centro/Sur","Exótica (Chile)","Medicinal / Ornamental","Jardín mediterráneo, nativo-andino","Hojas aromáticas; uso hepático medicinal clásico","Blanco","No"],
    [499,"Maqui","Aristotelia chilensis","Frutal / Medicinal","Exterior","Alta/Media","Moderado","Primavera","Perenne","2-5","2-3","Primavera","Sur","Nativa (Patagonia)","Frutal nativo / Ornamental","Jardín patagónico, borde","Frutos morados comestibles; muy antioxidante","Blanco","Sí"],
    [500,"Aguaribay / Molle","Schinus molle","Árbol frutal / Ornamental","Exterior","Alta","Escaso","Primavera","Perenne","5-12","5-10","Otoño","Centro/Norte","Nativa","Sombra / Ornamental / Nativo","Jardín nativo, xeriscape","Muy resistente; frutos rojos ornamentales; aromático","Amarillo/Verde","No"],
    # Más frutales y aromáticas para completar lote
    [501,"Tucumano (guayabo del país)","Acca sellowiana","Frutal","Exterior","Alta","Moderado","Primavera","Perenne","2-5","2-4","Primavera","Centro/Norte","Nativa","Frutal nativo / Ornamental","Jardín nativo, seto","Fruto con sabor tropical; flores bicolores ornamentales","Rojo/Blanco","Sí"],
    [502,"Guindo","Prunus cerasus","Frutal","Exterior","Alta","Moderado","Primavera","Caducifolio","3-6","3-5","Otoño","Sur","Exótica","Frutal / Ornamental","Jardín rural, huerto","Similar al cerezo pero más pequeño; fruto ácido","Blanco/Rosa","No"],
    [503,"Albaricoque / Damasco","Prunus armeniaca","Frutal","Exterior","Alta","Moderado","Primavera","Caducifolio","3-6","3-5","Otoño","Centro/Sur","Exótica","Frutal / Ornamental","Huerto familiar, jardín","Floración temprana rosada; fruto veraniego","Blanco/Rosa","No"],
    [504,"Fresón silvestre","Fragaria vesca","Frutal","Exterior","Alta/Media","Moderado","Primavera-Verano","Perenne","0.15-0.25","0.3","Otoño","Centro/Sur","Exótica/Naturalizada","Tapizante frutal / Borde","Jardín naturalista, bajo árboles","Fruto pequeño y muy aromático; tapizante","Blanco","Sí"],
    [505,"Lúcuma","Pouteria lucuma","Frutal","Exterior","Alta","Moderado","Primavera","Perenne","4-10","3-6","Primavera","Norte","Exótica (Andino)","Frutal / Ornamental","Jardín subtropical","Fruto andino de valor nutritivo; no tolera heladas","Amarillo/Verde","No"],
    [506,"Chirimoya","Annona cherimola","Frutal","Exterior","Alta","Moderado","Primavera-Verano","Caducifolio","5-10","4-7","Primavera","Norte","Exótica","Frutal / Sombra","Jardín subtropical","Fruto exótico de sabor cremoso; no soporta frío","Crema/Amarillo","No"],
    [507,"Carambola","Averrhoa carambola","Frutal","Exterior","Alta","Moderado","Todo el año","Perenne","4-8","3-6","Primavera","Norte","Exótica","Frutal / Ornamental","Jardín tropical","Fruto estrellado ornamental; requiere clima cálido","Rosa/Lila","No"],
    [508,"Guayabo","Psidium guajava","Frutal","Exterior","Alta","Moderado","Primavera-Verano","Perenne","3-8","3-6","Primavera","Norte","Exótica","Frutal / Sombra","Jardín subtropical","Rústico; fruto rico en vitamina C; corteza ornamental","Blanco","No"],
    [509,"Pitanga","Eugenia uniflora","Frutal","Exterior","Alta/Media","Moderado","Primavera","Perenne","3-6","2-4","Primavera","Norte","Exótica (Brasil)","Frutal / Seto / Ornamental","Jardín subtropical, seto","Fruto rojo-anaranjado aromático; hojas rojizas en brote","Blanco","Sí"],
    [510,"Jabuticaba","Plinia cauliflora","Frutal","Exterior","Alta/Media","Moderado","Primavera-Verano","Perenne","3-8","3-6","Primavera","Norte","Exótica (Brasil)","Frutal exótico / Ornamental","Jardín tropical","Único: fruto nace en el tronco; muy ornamental","Blanco","No"],
    [511,"Murta / Arrayán","Ugni molinae","Frutal","Exterior","Alta/Media","Moderado","Primavera","Perenne","0.5-2","0.5-1.5","Primavera","Sur","Nativa (Patagonia)","Frutal nativo / Seto / Ornamental","Jardín patagónico, seto","Fruto muy aromático; uso en licores y mermeladas","Blanco/Rosa","Sí"],
    [512,"Saúco negro","Sambucus nigra","Frutal / Medicinal","Exterior","Alta/Media","Moderado","Primavera","Caducifolio","2-6","2-4","Otoño","Centro/Sur","Exótica","Frutal / Medicinal / Naturalista","Jardín naturalista, borde de agua","Flores y frutos comestibles; medicinal; atrae fauna","Blanco/Crema","No"],
    [513,"Rosa mosqueta","Rosa rubiginosa","Frutal / Ornamental","Exterior","Alta","Escaso","Primavera-Verano","Caducifolio","1-2.5","1-2","Primavera","Centro/Sur","Exótica/Naturalizada","Frutal / Ornamental / Cerco","Jardín rural, seto espinoso","Escaramujos ricos en vitamina C; invasiva en Patagonia","Rosa","No"],
    [514,"Tomate cherry","Solanum lycopersicum var.","Frutal/Hortaliza","Exterior","Alta","Moderado","Verano","Anual","0.5-1.5","0.5","Primavera","Todo el país","Exótica","Huerto ornamental / Maceta","Huerto, maceta, jardín comestible","Frutos ornamentales coloridos; fácil en maceta","Amarillo","Sí"],
    [515,"Pimiento ornamental","Capsicum annuum","Frutal/Hortaliza","Exterior/Interior","Alta","Moderado","Verano","Anual","0.3-0.6","0.3-0.5","Primavera","Todo el país","Exótica","Ornamental / Huerto / Maceta","Maceta, jardín comestible, interior","Frutos coloridos muy ornamentales (rojo, amarillo, violeta)","Blanco","Sí"],
    [516,"Moringa","Moringa oleifera","Árbol / Medicinal","Exterior","Alta","Escaso","Variable","Caducifolio","4-12","3-6","Primavera","Norte","Exótica","Medicinal / Ornamental","Jardín subtropical, diseño comestible","Árbol 'milagroso'; hojas muy nutritivas; crece rápido","Blanco/Crema","No"],
    [517,"Estevia","Stevia rebaudiana","Aromática / Medicinal","Exterior","Alta","Moderado","Verano-Otoño","Perenne","0.4-0.8","0.3-0.5","Primavera","Centro/Norte","Exótica","Medicinal / Huerto / Edulcorante","Maceta, huerto, jardín comestible","Hojas 300x más dulces que el azúcar; uso medicinal","Blanco/Lila","Sí"],
    [518,"Yerba mate","Ilex paraguariensis","Árbol / Cultural","Exterior","Media (semisombra)","Moderado","Primavera","Perenne","4-15","3-6","Primavera","Norte (Misiones)","Nativa","Frutal nativo / Cultural / Sombra","Jardín misionero, subtropical","Cultivo icónico argentino; requiere clima subtropical húmedo","Blanco","No"],
    [519,"Tilo","Tilia cordata","Aromática / Árbol","Exterior","Alta","Moderado","Verano","Caducifolio","10-20","6-12","Otoño","Centro/Sur","Exótica","Sombra / Aromática / Medicinal","Parque, jardín grande","Flor aromática para infusión sedante; árbol majestuoso","Amarillo/Blanco","No"],
    [520,"Verbena","Verbena officinalis","Aromática / Medicinal","Exterior","Alta","Moderado","Primavera-Otoño","Perenne","0.3-0.6","0.3-0.5","Primavera","Todo el país","Exótica","Medicinal / Borde","Borde, jardín naturalista","Uso medicinal antiguo; atrae polinizadores","Lila/Rosa","Sí"],
    [521,"Árnica","Arnica montana","Medicinal","Exterior","Alta","Moderado","Verano","Perenne","0.3-0.6","0.3-0.4","Otoño/Primavera","Sur","Exótica","Medicinal / Borde ornamental","Jardín naturalista, montañoso","Flores amarillas; uso medicinal tópico; planta de altura","Amarillo","Sí"],
    [522,"Borraja","Borago officinalis","Aromática / Ornamental","Exterior","Alta","Moderado","Primavera-Verano","Anual","0.3-0.7","0.3-0.4","Otoño/Primavera","Todo el país","Exótica","Medicinal / Borde / Huerto","Huerto, borde, jardín naturalista","Flores azul estrelladas comestibles y ornamentales","Azul","Sí"],
    [523,"Gordolobo","Verbascum thapsus","Medicinal","Exterior","Alta","Escaso","Verano","Bienal","1-2","0.3-0.5","Otoño","Centro/Sur","Exótica","Medicinal / Estructural","Jardín naturalista, borde alto","Espiga alta muy ornamental; hojas lanosas; medicinal","Amarillo","No"],
    [524,"Milenrama","Achillea millefolium","Medicinal / Ornamental","Exterior","Alta","Escaso","Verano","Perenne","0.4-0.8","0.4-0.6","Primavera","Todo el país","Exótica/Naturalizada","Medicinal / Borde / Naturalista","Jardín naturalista, pradera","Flores en corimbo blancas/rosadas; muy resistente","Blanco/Rosa/Rojo","Sí"],
    [525,"Equinácea","Echinacea purpurea","Medicinal / Ornamental","Exterior","Alta","Moderado","Verano-Otoño","Perenne","0.6-1","0.4-0.6","Primavera","Centro/Sur","Exótica","Medicinal / Borde ornamental","Jardín naturalista, borde","Flores vistosas; inmunoestimulante; atrae mariposas","Rosa/Violeta","Sí"],
    [526,"Ajo","Allium sativum","Aromática / Hortaliza","Exterior","Alta","Moderado","Primavera","Perenne (bulbo)","0.3-0.6","0.1","Otoño","Todo el país","Exótica","Repelente / Huerto compañero","Huerto, borde","Repele plagas; cultivo fácil; bulbo en verano","Blanco","Sí"],
    [527,"Cebollín de río","Allium ampeloprasum","Aromática","Exterior","Alta","Moderado","Primavera","Perenne","0.5-1","0.2","Otoño","Centro/Norte","Exótica","Aromática / Huerto","Borde, huerto","Puerro silvestre; hoja más gruesa que ciboulette","Violeta/Rosa","Sí"],
    [528,"Mejorana","Origanum majorana","Aromática","Exterior","Alta","Moderado","Verano","Perenne (anual en frío)","0.3-0.5","0.2-0.4","Primavera","Centro/Norte","Exótica","Aromática / Huerto","Huerto, maceta, jardín mediterráneo","Similar al orégano pero más dulce; uso culinario","Blanco/Rosa","Sí"],
    [529,"Apio","Apium graveolens","Aromática / Hortaliza","Exterior","Alta/Media","Frecuente","Verano","Bienal","0.4-0.8","0.3","Otoño/Primavera","Todo el país","Exótica","Aromática / Huerto","Huerto, maceta","Requiere humedad constante; semillas aromáticas","Blanco","Sí"],
    [530,"Angélica","Angelica archangelica","Aromática / Medicinal","Exterior","Alta/Media","Moderado","Verano","Bienal","1-2.5","0.5-1","Otoño","Centro/Sur","Exótica","Medicinal / Estructural","Jardín naturalista, borde húmedo","Planta bienal imponente; uso en licores y medicina","Blanco/Crema","No"],
    [531,"Laurel de cocina","Laurus nobilis","Aromática / Árbol","Exterior","Alta/Media","Moderado","Primavera","Perenne","3-8","2-5","Primavera","Centro/Norte","Exótica","Aromática / Estructural / Seto","Jardín mediterráneo, seto formal, maceta","Hojas perennes aromáticas; muy versátil en paisajismo","Amarillo/Crema","Sí"],
    [532,"Azafrán","Crocus sativus","Aromática / Ornamental","Exterior","Alta","Escaso","Otoño","Bulbo","0.1-0.15","0.1","Verano","Centro","Exótica","Aromática / Ornamental","Jardín de bulbos, borde","Especia más cara del mundo; florece en otoño; bulbo verano","Violeta","Sí"],
    [533,"Cártamo","Carthamus tinctorius","Aromática / Tintórea","Exterior","Alta","Escaso","Verano","Anual","0.5-1","0.3-0.5","Primavera","Centro/Norte","Exótica","Ornamental / Tintórea","Jardín naturalista, corte","Flores naranjas-amarillas para secar; colorante natural","Naranja/Amarillo","No"],
    [534,"Tagetes / Copete","Tagetes erecta","Aromática / Ornamental","Exterior","Alta","Moderado","Verano-Otoño","Anual","0.3-0.8","0.3-0.5","Primavera","Todo el país","Exótica","Repelente / Borde / Huerto","Huerto compañero, borde, maceta","Repele nematodos y pulgones; muy fácil cultivo","Naranja/Amarillo","Sí"],
    [535,"Catnip / Nébeda","Nepeta cataria","Aromática","Exterior","Alta","Escaso","Verano","Perenne","0.3-0.8","0.3-0.6","Primavera","Centro/Sur","Exótica","Aromática / Repelente / Borde","Jardín mediterráneo, borde","Atrae gatos; repele pulgones; flores lila ornamentales","Blanco/Lila","Sí"],
    [536,"Lavanda francesa","Lavandula stoechas","Aromática / Ornamental","Exterior","Alta","Escaso","Primavera-Verano","Perenne","0.4-0.7","0.4-0.6","Primavera","Centro/Norte","Exótica","Aromática / Borde / Ornamental","Jardín mediterráneo, maceta","Flor con penacho ornamental; más sensible al calor húmedo","Violeta/Morado","Sí"],
    [537,"Verbena limón","Aloysia triphylla","Aromática","Exterior","Alta","Moderado","Verano","Caducifolio","1-3","1-2","Primavera","Centro/Norte","Exótica","Aromática / Medicinal","Jardín mediterráneo, maceta grande","Aroma cítrico intensísimo; hojas para infusión","Blanco/Lila","Sí"],
    [538,"Coriandro perenne","Eryngium foetidum","Aromática","Exterior/Interior","Alta/Media","Moderado","Todo el año","Perenne","0.2-0.5","0.3","Primavera","Norte","Exótica","Aromática / Huerto tropical","Maceta, huerto tropical","Sustituto del cilantro en climas cálidos; no bolt","Verde/Blanco","Sí"],
    [539,"Isopo azul","Hyssopus officinalis 'Blue'","Aromática","Exterior","Alta","Escaso","Verano","Perenne","0.4-0.6","0.3-0.5","Primavera","Centro/Sur","Exótica","Aromática / Borde / Medicinal","Jardín mediterráneo, borde","Cultivar azul intenso; atrae mariposas y abejas","Azul intenso","Sí"],
    [540,"Matricaria / Manzanilla romana","Chamaemelum nobile","Aromática / Tapizante","Exterior","Alta","Moderado","Primavera-Verano","Perenne","0.1-0.3","0.3-0.5","Otoño/Primavera","Centro/Sur","Exótica","Aromática / Tapizante","Tapizante aromático, borde, jardín de senderos","Suelo perfumado al pisarlo; tapizante entre piedras","Blanco/Amarillo","Sí"],
    [541,"Regaliz","Glycyrrhiza glabra","Aromática / Medicinal","Exterior","Alta","Moderado","Verano","Perenne","0.8-1.5","0.5-1","Primavera","Centro","Exótica","Medicinal / Ornamental","Jardín medicinal","Raíz de sabor dulce; flores lila ornamentales","Violeta/Lila","No"],
    [542,"Hinojo de mar","Crithmum maritimum","Aromática","Exterior","Alta","Escaso","Verano","Perenne","0.3-0.5","0.3-0.5","Primavera","Centro (costas)","Exótica","Aromática / Costera / Xeriscape","Jardín costero, xeriscape","Muy tolerante a sal; hojas crujientes comestibles","Blanco/Amarillo","Sí"],
    [543,"Abeto de limón","Eucalyptus citriodora","Aromático / Árbol","Exterior","Alta","Escaso","Primavera","Perenne","15-30","5-10","Primavera","Centro/Norte","Exótica","Aromática / Sombra / Estructural","Parque grande, forestación","Corteza blanca ornamental; aroma a limón; crece rápido","Blanco/Crema","No"],
    [544,"Mirto","Myrtus communis","Aromática / Ornamental","Exterior","Alta","Escaso","Primavera-Verano","Perenne","1-3","1-2","Primavera","Centro/Norte","Exótica","Aromática / Seto / Tópico","Jardín mediterráneo, seto","Bayas azul-negro comestibles; hojas aromáticas; fruto para licor","Blanco","Sí"],
    [545,"Arrayán chileno","Luma apiculata","Aromática / Árbol","Exterior","Alta/Media","Moderado","Verano","Perenne","5-15","3-7","Primavera","Sur (Patagonia)","Nativa","Ornamental / Aromática / Nativo","Jardín patagónico, borde de agua","Corteza naranja-canela exfoliante muy ornamental; nativo","Blanco","No"],
    [546,"Lavanda de mar","Limonium latifolium","Ornamental / Aromática","Exterior","Alta","Escaso","Verano","Perenne","0.3-0.6","0.4-0.6","Primavera","Centro/Norte","Exótica","Borde / Xeriscape / Costera","Jardín xeriscape, costero, borde","Flores para secar; tolerante a sal; muy decorativa","Lila/Violeta","Sí"],
    [547,"Palo azul","Cyclolepis genistoides","Medicinal / Ornamental","Exterior","Alta","Escaso","Verano","Perenne","0.5-1.5","0.5-1","Primavera","Centro/Norte","Nativa","Medicinal / Xeriscape / Nativo","Jardín nativo xérico","Uso medicinal urinario tradicional; muy resistente sequía","Lila","No"],
    [548,"Anís del campo","Tagetes minuta","Aromática","Exterior","Alta","Escaso","Verano-Otoño","Anual","1-2","0.3-0.5","Primavera","Centro/Norte","Nativa","Aromática / Repelente","Huerto, jardín nativo","Nativa argentina; repele nematodos; aroma anisado intenso","Amarillo","No"],
    [549,"Oreganillo","Lippia graveolens","Aromática","Exterior","Alta","Escaso","Primavera-Verano","Perenne","0.5-1.5","0.5-1","Primavera","Norte","Nativa","Aromática / Medicinal","Jardín nativo norteño","Orégano nativo del norte; muy aromático; calor y sequía","Blanco/Amarillo","Sí"],
    [550,"Muña muña","Minthostachys setosa","Aromática","Exterior","Alta","Escaso","Verano","Perenne","0.3-0.8","0.3-0.5","Primavera","Noroeste","Nativa (Andes)","Aromática / Medicinal andina","Jardín andino, xeriscaping","Aromática andina; uso medicinal digestivo; muy resistente","Blanco","Sí"],
]

# Insert Lote 2 rows starting at row 439
print(f"Adding Lote 2 ({len(lote2)} plants)...")
for plant_data in lote2:
    ws.append(plant_data)

# Apply formatting to all new rows (282 to end)
total_rows = ws.max_row
print(f"Applying formatting to rows 282-{total_rows}...")
for r in range(282, total_rows + 1):
    fill = FILL_ODD if (r % 2 == 0) else FILL_EVEN
    row = ws[r]
    for i, cell in enumerate(row, start=1):
        cell.fill = fill
        cell.border = BORDER
        cell.font = FONT
        cell.alignment = ALIGN_LEFT if i in LEFT_COLS else ALIGN

# Fix row heights for new rows
for r in range(282, total_rows + 1):
    ws.row_dimensions[r].height = 40

wb.save('flora_argentina_completa_v5.xlsx')
print(f"Saved flora_argentina_completa_v5.xlsx with {total_rows} data rows + header")
EOF